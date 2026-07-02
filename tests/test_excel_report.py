"""Tests de excel_report.py : on recharge le .xlsx produit et on vérifie le contenu."""
from datetime import date
from io import BytesIO

import openpyxl

import app
import excel_report


def _projet():
    return {"no": "12345", "id_project": 7, "semaine": date(2026, 6, 21),
            "adresse": "123 rue Principale", "lat": None, "lon": None}


def _day_rempli():
    q = app._empty_quart()
    q["personnel"] = ["Mathis Lajeunesse"]
    q["temp_am"] = 12.0
    q["conditions"] = ["Ensoleillé"]
    q["heures"] = {"Mathis Lajeunesse": {"Excavation": {"TR": 8.0, "TS": 1.0}}}
    q["prime_codes"] = {"Mathis Lajeunesse": ["S"]}
    return {"date": date(2026, 6, 22), "quarts": {"Jour": q}}


def _day_vide():
    return {"date": date(2026, 6, 22), "quarts": {"Jour": app._empty_quart()}}


def _all_text(ws):
    return "\n".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value not in (None, ""))


def test_build_day_workbook_une_feuille_et_entete():
    buf = excel_report.build_day_workbook(_projet(), "Lundi", _day_rempli(), "Test User")
    wb = openpyxl.load_workbook(buf)
    assert wb.sheetnames == ["Lundi"]
    txt = _all_text(wb["Lundi"])
    assert "Rapport journalier" in txt
    assert "12345" in txt                      # No Projet (bandeau titre)
    assert "Contremaître :" in txt             # champ méta (remplace l'adresse)
    assert "Mathis Lajeunesse" in txt          # ligne de personnel
    assert "Exporté par Test User" in txt      # estampille


def test_build_day_workbook_largeurs_colonnes_ajustees():
    # Garde-fou : la colonne Nom/Activité et la colonne Commentaire sont larges,
    # plusieurs colonnes ont une largeur (pas seulement A).
    buf = excel_report.build_day_workbook(_projet(), "Lundi", _day_rempli(), "")
    ws = openpyxl.load_workbook(buf)["Lundi"]
    larges = {k: d.width for k, d in ws.column_dimensions.items() if d.width}
    assert "A" in larges and larges["A"] >= 30          # colonne Nom / Activité
    assert max(larges.values()) >= 30                    # colonne Commentaire
    assert len(larges) >= 5                              # plusieurs colonnes réglées


def test_build_day_workbook_groupe_par_employe():
    # Option B : ligne du nom, sous-ligne d'activité, ligne Total.
    buf = excel_report.build_day_workbook(_projet(), "Lundi", _day_rempli(), "")
    ws = openpyxl.load_workbook(buf)["Lundi"]
    col_a = [str(c.value) for c in ws["A"] if c.value not in (None, "")]
    assert any("Mathis Lajeunesse" in v for v in col_a)        # ligne nom
    assert any("Excavation" in v for v in col_a)               # sous-ligne activité
    assert any(v.strip() == "Total" for v in col_a)            # ligne Total


def test_build_day_workbook_responsable_repli_sur_exportateur():
    # Quart sans responsable estampillé : le nom de l'exportateur (personne
    # connectée) doit apparaître comme Contremaître dans le bloc méta.
    q = app._empty_quart()
    q["personnel"] = ["Bob"]
    q["heures"] = {"Bob": {"Excavation": {"TR": 8.0, "TS": 0.0}}}
    assert q["responsable"] == ""
    day = {"date": date(2026, 6, 22), "quarts": {"Jour": q}}
    buf = excel_report.build_day_workbook(_projet(), "Lundi", day, "Marie-Pier Arsenault")
    txt = _all_text(openpyxl.load_workbook(buf)["Lundi"])
    assert "Contremaître :" in txt
    assert "Marie-Pier Arsenault" in txt        # repli sur l'exportateur


def test_build_day_workbook_temperature_et_conditions():
    buf = excel_report.build_day_workbook(_projet(), "Lundi", _day_rempli(), "")
    txt = _all_text(openpyxl.load_workbook(buf)["Lundi"])
    assert "Température" in txt
    assert "Ensoleillé" in txt        # condition saisie (_day_rempli)
    assert "12" in txt                # temp_am = 12.0


def test_build_day_workbook_meteo_dans_panneau_entete():
    # La météo va dans le panneau d'en-tête (haut à droite), pas sous le quart :
    # le libellé « Température ext. » est celui du panneau, en lignes 4-5, cols E+.
    ws = openpyxl.load_workbook(
        excel_report.build_day_workbook(_projet(), "Lundi", _day_rempli(), ""))["Lundi"]
    positions = [(c.row, c.column) for row in ws.iter_rows() for c in row
                 if c.value == "Température ext."]
    assert positions, "libellé du panneau météo absent"
    r, col = positions[0]
    assert r in (4, 5)                 # zone méta (haut de feuille)
    assert col >= 5                    # colonne E ou plus (côté droit)


def test_build_day_workbook_sans_quadrillage():
    ws = openpyxl.load_workbook(
        excel_report.build_day_workbook(_projet(), "Lundi", _day_rempli(), ""))["Lundi"]
    assert ws.sheet_view.showGridLines is False


def test_build_day_workbook_entete_travaux_effectues():
    buf = excel_report.build_day_workbook(_projet(), "Lundi", _day_rempli(), "")
    txt = _all_text(openpyxl.load_workbook(buf)["Lundi"])
    assert "Travaux effectués" in txt


def test_build_day_workbook_plages_horaires_par_activite():
    # Une activité en mode « plage » affiche ses créneaux début–fin et leur type.
    q = app._empty_quart()
    q["personnel"] = ["Bob"]
    q["heures"] = {"Bob": {"Excavation": {
        "mode": "plage",
        "ranges": [{"debut": "07:00", "fin": "12:00", "type": "TR"},
                   {"debut": "13:00", "fin": "14:00", "type": "TS"}],
        "TR": 5.0, "TS": 1.0}}}
    day = {"date": date(2026, 6, 22), "quarts": {"Jour": q}}
    buf = excel_report.build_day_workbook(_projet(), "Lundi", day, "")
    ws = openpyxl.load_workbook(buf)["Lundi"]
    col_a = [str(c.value) for c in ws["A"] if c.value not in (None, "")]
    assert any("07:00 – 12:00" in v for v in col_a)         # plage 1
    assert any("13:00 – 14:00" in v for v in col_a)         # plage 2
    vals = [c.value for row in ws.iter_rows() for c in row]
    assert 5.0 in vals and 1.0 in vals                       # heures TR / TS dérivées


def test_build_day_workbook_heures_et_prime_presentes():
    buf = excel_report.build_day_workbook(_projet(), "Lundi", _day_rempli(), "")
    wb = openpyxl.load_workbook(buf)
    vals = [c.value for row in wb["Lundi"].iter_rows() for c in row]
    assert 8.0 in vals and 1.0 in vals     # TR et TS
    ws = wb["Lundi"]
    total = next(r for r in range(1, ws.max_row + 1)
                 if ws.cell(r, 1).value == "Total")
    assert ws.cell(total, 7).value == "S"   # code de prime dans la colonne Prime


def test_build_day_workbook_jour_vide_sans_personnel():
    buf = excel_report.build_day_workbook(_projet(), "Lundi", _day_vide(), "")
    wb = openpyxl.load_workbook(buf)
    txt = _all_text(wb["Lundi"])
    assert "Rapport journalier" in txt
    assert "Mathis" not in txt


def test_build_day_workbook_quart_et_legende_prime():
    # Le quart est un champ du bloc méta (plus de bandeau) et la légende des
    # codes de prime est présente.
    txt = _all_text(openpyxl.load_workbook(
        excel_report.build_day_workbook(_projet(), "Lundi", _day_rempli(), ""))["Lundi"])
    assert "Quart :" in txt
    assert "Code de prime" in txt


def test_build_day_workbook_ordre_colonnes_code_puis_hrs_eq():
    # En-tête du tableau : « Code Éq. » (col E) précède « Hrs Éq. » (col F).
    ws = openpyxl.load_workbook(
        excel_report.build_day_workbook(_projet(), "Lundi", _day_rempli(), ""))["Lundi"]
    header = next(r for r in range(1, ws.max_row + 1)
                  if ws.cell(r, 1).value == "Nom / Activité")
    assert ws.cell(header, 5).value == "Code Éq."
    assert ws.cell(header, 6).value == "Hrs Éq."


def test_build_week_workbook_une_feuille_par_jour_rempli():
    jours = {j: _day_vide() for j in app.JOURS}
    jours["Lundi"] = _day_rempli()
    jours["Mercredi"] = _day_rempli()
    buf = excel_report.build_week_workbook(_projet(), jours, app.JOURS, "")
    wb = openpyxl.load_workbook(buf)
    assert wb.sheetnames == ["Lundi", "Mercredi"]


def test_build_week_workbook_feuilles_modernisees():
    jours = {j: _day_vide() for j in app.JOURS}
    jours["Lundi"] = _day_rempli()
    buf = excel_report.build_week_workbook(_projet(), jours, app.JOURS, "")
    txt = _all_text(openpyxl.load_workbook(buf)["Lundi"])
    assert "Rapport journalier" in txt
    assert "Total de la journée" in txt
    assert "Revu par" in txt


def test_build_day_workbook_total_de_la_journee():
    q = app._empty_quart()
    q["personnel"] = ["A", "B"]
    q["heures"] = {"A": {"Excavation": {"TR": 4.0, "TS": 0.0}},
                   "B": {"Excavation": {"TR": 4.0, "TS": 1.0}}}
    day = {"date": date(2026, 6, 22), "quarts": {"Jour": q}}
    ws = openpyxl.load_workbook(
        excel_report.build_day_workbook(_projet(), "Lundi", day, ""))["Lundi"]
    # Trouver la ligne « Total de la journée » et vérifier ses totaux TR/TS.
    row = next(r for r in range(1, ws.max_row + 1)
               if ws.cell(r, 1).value == "Total de la journée")
    assert ws.cell(row, 3).value == 8.0    # TR : 4 + 4
    assert ws.cell(row, 4).value == 1.0    # TS : 0 + 1


def test_build_day_workbook_bas_de_page():
    buf = excel_report.build_day_workbook(_projet(), "Lundi", _day_rempli(), "")
    txt = _all_text(openpyxl.load_workbook(buf)["Lundi"])
    assert "Codes d'équipement" in txt
    assert "Commentaires / plaintes / suggestions" in txt
    assert "Revu par" in txt
    assert "Approuvé par" in txt


def test_build_day_email_renvoie_sujet_nom_et_bytes():
    subject, html, filename, data = excel_report.build_day_email(
        _projet(), "Lundi", _day_rempli(), "Test User")
    assert "12345" in subject and "Lundi" in subject
    assert filename == "Rapport_12345_2026-06-22.xlsx"
    assert isinstance(data, bytes) and data[:2] == b"PK"   # signature zip/xlsx
    assert "Test User" in html
