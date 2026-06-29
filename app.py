"""
Rapport Journalier Ondel — application Streamlit
Optimisée pour Tablette (UX Dashboard + Géolocalisation)
"""

import json
import os
import urllib.parse
import urllib.request
import base64
import datetime
from datetime import date, timedelta
from io import BytesIO

import pandas as pd
import streamlit as st
import data_source
import reports

try:
    from streamlit_js_eval import get_geolocation
except Exception:  # composant optionnel — l'app marche sans (pas de capture GPS)
    get_geolocation = None
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Constantes / configuration
# --------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(BASE_DIR, "ondel.png")

JOURS = ["Dimanche", "Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"]
# Mois en français (indexés sur d.month, 1-12) — évite la dépendance à la locale
# système qui rendrait strftime("%B") en anglais.
MOIS = ["", "janvier", "février", "mars", "avril", "mai", "juin",
        "juillet", "août", "septembre", "octobre", "novembre", "décembre"]
MOIS_ABBR = ["", "janv.", "févr.", "mars", "avr.", "mai", "juin",
             "juil.", "août", "sept.", "oct.", "nov.", "déc."]

def fr_date_long(d):   # ex. "18 juin 2026"
    return f"{d.day:02d} {MOIS[d.month]} {d.year}"

def fr_date_short(d):  # ex. "18 juin"
    return f"{d.day:02d} {MOIS_ABBR[d.month]}"
QUARTS = ["", "Jour", "Soir", "Nuit"]
QUART_NAMES = [q for q in QUARTS if q]  # ["Jour", "Soir", "Nuit"], ordonné
CONDITIONS = [
    "Ensoleillé", "Partiellement nuageux", "Nuageux", "Couvert",
    "Pluie", "Averses", "Neige", "Verglas", "Brouillard", "Venteux",
]

ACT_KEYS = [f"h{i}" for i in range(8)]
AUTRE_KEYS = [f"a{i}" for i in range(4)]
HOUR_KEYS = ACT_KEYS + AUTRE_KEYS

EQUIP_CODES = [
    ("C", "Camion"), ("N", "Nacelle"), ("É", "Éch. Hyd."),
    ("D", "Détecteur"), ("G", "Grue"), ("BT", "Chariot élév."),
]
EQUIP_CODE_VALUES = [c for c, _ in EQUIP_CODES]
_EQUIP_CODE_LABELS = dict(EQUIP_CODES)


def _equip_code_label(code):
    return f"{code} — {_EQUIP_CODE_LABELS.get(code, code)}"


# Couleurs Ondel
ONDEL_GREEN = "#0999AA"
ONDEL_GREEN_DARK = "#077A88"
ONDEL_ACCENT = "#26313F"
ONDEL_ACCENT_DK = "#1A222C"
ONDEL_LIGHT_BG = "#F8FAFB"

# Styles Excel
_TEAL = "0999AA"
_TEAL_DK = "077A88"
_BAND = "EEF8F9"
_GREY = "D9E2E4"
_THIN = Side(style="thin", color=_GREY)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_F_TITLE = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
_F_HEAD = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_F_LABEL = Font(name="Calibri", size=10, bold=True, color=_TEAL_DK)
_F_TOTAL = Font(name="Calibri", size=10, bold=True, color="0E2A2E")
_FILL_TITLE = PatternFill("solid", fgColor=_TEAL)
_FILL_HEAD = PatternFill("solid", fgColor=_TEAL_DK)
_FILL_BAND = PatternFill("solid", fgColor=_BAND)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_HOURS_FMT = "0.00"

# --------------------------------------------------------------------------
# State
# --------------------------------------------------------------------------
def init_state():
    if "projet" not in st.session_state:
        st.session_state.projet = {"no": "", "id_project": None, "semaine": _week_start(date.today()), "adresse": "", "lat": None, "lon": None}
    if "jours" not in st.session_state:
        st.session_state.jours = {j: _empty_day() for j in JOURS}
        _apply_week_dates(st.session_state.projet["semaine"])
    if "view" not in st.session_state:
        st.session_state.view = "dashboard"
    if "active_day" not in st.session_state:
        st.session_state.active_day = None
    if "dirty" not in st.session_state:
        st.session_state.dirty = False
    if "loaded_key" not in st.session_state:
        st.session_state.loaded_key = None
    if "schema_ready" not in st.session_state:
        st.session_state.schema_ready = False

def current_user():
    """Identité connectée : {"name", "email"}. Chaînes vides si non connecté."""
    user = getattr(st, "user", None)
    if user is None or not getattr(user, "is_logged_in", False):
        return {"name": "", "email": ""}
    return {
        "name": getattr(user, "name", "") or "",
        "email": getattr(user, "email", "") or "",
    }

# --------------------------------------------------------------------------
# Persistance Neon (chargement / sauvegarde des rapports saisis)
# --------------------------------------------------------------------------
def _report_key():
    """Clé métier courante (id_project, semaine), ou None si projet non choisi."""
    proj = st.session_state.projet
    idp = proj.get("id_project")
    wk = proj.get("semaine")
    if idp is None or not isinstance(wk, date):
        return None
    return (int(idp), wk)


def _clear_quart_widget_state(jour, quart_name):
    prefixes = (f"tr_{jour}_{quart_name}_", f"ts_{jour}_{quart_name}_",
                f"eqc_{jour}_{quart_name}_", f"eqh_{jour}_{quart_name}_",
                f"p_{jour}_{quart_name}_", f"c_{jour}_{quart_name}_",
                f"acts_{jour}_{quart_name}",
                f"cond_pills_{jour}_{quart_name}",
                f"{jour}_{quart_name}_temp_am", f"{jour}_{quart_name}_temp_pm",
                f"{jour}_{quart_name}_cond", f"resource_sel_{jour}_{quart_name}",
                f"personnel_pills_{jour}_{quart_name}", f"note_{jour}_{quart_name}",
                f"show_geoloc_{jour}_{quart_name}", f"geoloc_{jour}_{quart_name}",
                f"ranges_{jour}_{quart_name}_", f"rangeseq_{jour}_{quart_name}_",
                f"mode_{jour}_{quart_name}_", f"rg_deb_{jour}_{quart_name}_",
                f"rg_fin_{jour}_{quart_name}_", f"rg_knd_{jour}_{quart_name}_")
    for k in list(st.session_state.keys()):
        if any(k.startswith(p) for p in prefixes):
            del st.session_state[k]


def _mark_dirty():
    st.session_state.dirty = True


def load_report_into_state():
    """Charge le rapport (projet+semaine) depuis Neon dans le state, une seule
    fois par clé. Si aucun rapport enregistré, repart d'une semaine vierge."""
    key = _report_key()
    if key is None or st.session_state.loaded_key == key:
        return
    idp, wk = key
    try:
        data = reports.load_report(idp, wk)
    except Exception:
        data = None

    st.session_state.jours = {j: _empty_day() for j in JOURS}
    _apply_week_dates(wk)

    if data:
        for jour in JOURS:
            day = st.session_state.jours[jour]
            saved = data["days_by_date"].get(day["date"])
            if saved and saved.get("quarts"):
                day["quarts"] = {}
                for qname, q in saved["quarts"].items():
                    nq = _empty_quart()
                    if q:
                        nq.update(q)
                    day["quarts"][qname] = nq
                if not day["quarts"]:
                    day["quarts"] = {"Jour": _empty_quart()}

    # Purge des clés de widgets de tous les jours/quarts
    for jour in JOURS:
        for qname in _day_quart_names(st.session_state.jours[jour]):
            _clear_quart_widget_state(jour, qname)
        st.session_state.pop(f"active_quart_{jour}", None)
    st.session_state.pop("team_pers", None)
    st.session_state.pop("team_equip", None)

    st.session_state.loaded_key = key
    st.session_state.dirty = False


def save_report_from_state():
    """Persiste le state courant vers Neon. Renvoie (ok, message)."""
    try:
        user = current_user()
        # Responsable verrouillé sur l'utilisateur connecté pour tous les quarts.
        for day in st.session_state.jours.values():
            for quart in day.get("quarts", {}).values():
                quart["responsable"] = user["name"]
        reports.save_report(
            st.session_state.projet, {},
            st.session_state.jours, JOURS,
            saved_by=user["email"],
        )
        st.session_state.dirty = False
        return True, "Rapport enregistré ✓"
    except Exception as exc:  # noqa: BLE001
        return False, f"Échec de l'enregistrement : {exc}"


def _empty_quart():
    return {
        "responsable": "", "activites": [], "autres": [],
        "personnel": [], "equipements": [],
        "temp_am": None, "temp_pm": None, "conditions": [],
        "heures": {}, "prime": {}, "commentaire_ligne": {},
        "equip_codes": {}, "equip_hours": {},
        "description": "",
    }

def _empty_day():
    return {"date": None, "quarts": {"Jour": _empty_quart()}}

def _day_quart_names(day):
    return [q for q in QUART_NAMES if q in day["quarts"]]

def _week_start(d):
    if not isinstance(d, date):
        return d
    return d - timedelta(days=(d.weekday() + 1) % 7)

def _apply_week_dates(semaine):
    start = _week_start(semaine)
    if not isinstance(start, date): return
    for i, jour in enumerate(JOURS):
        st.session_state.jours[jour]["date"] = start + timedelta(days=i)

# --------------------------------------------------------------------------
# Météo & Géo
# --------------------------------------------------------------------------
def _wmo_to_condition(code):
    c = int(code)
    if c in (0, 1): return "Ensoleillé"
    if c == 2: return "Partiellement nuageux"
    if c == 3: return "Couvert"
    if c in (45, 48): return "Brouillard"
    if c in (51, 53, 55, 61, 63, 65): return "Pluie"
    if c in (56, 57, 66, 67): return "Verglas"
    if c in (71, 73, 75, 77, 85, 86): return "Neige"
    if c in (80, 81, 82, 95, 96, 99): return "Averses"
    return "Nuageux"

def _codes_to_condition(codes):
    day = codes[6:19] if len(codes) >= 19 else list(codes)
    valid = [c for c in day if c is not None]
    return _wmo_to_condition(max(valid)) if valid else "Ensoleillé"

@st.cache_data(show_spinner=False, ttl=3600)
def _fetch_day_weather(lat, lon, d_iso):
    try:
        url = f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}&hourly=temperature_2m,weather_code&timezone=auto&start_date={d_iso}&end_date={d_iso}"
        with urllib.request.urlopen(url, timeout=10) as r:
            data = json.load(r)
        hourly = data.get("hourly", {})
        temps = hourly.get("temperature_2m", [])
        codes = hourly.get("weather_code", [])
        return {
            "temp_am": temps[9] if len(temps) > 9 else None,
            "temp_pm": temps[15] if len(temps) > 15 else None,
            "conditions": [_codes_to_condition(codes)],
        }
    except: return None

def _geocode_address(address):
    """Géocode une adresse via Nominatim (OpenStreetMap) -> (lat, lon) ou None."""
    try:
        import urllib.parse
        encoded = urllib.parse.quote(address)
        url = f"https://nominatim.openstreetmap.org/search?q={encoded}&format=json&limit=1"
        with urllib.request.urlopen(url, timeout=10) as r:
            results = json.load(r)
        if results:
            return float(results[0]["lat"]), float(results[0]["lon"])
    except:
        pass
    return None

def _fill_weather_for_quart(jour_name, quart_name):
    proj = st.session_state.projet
    day = st.session_state.jours[jour_name]
    quart = day["quarts"][quart_name]

    if (not proj.get("lat") or not proj.get("lon")) and proj.get("adresse"):
        coords = _geocode_address(proj["adresse"])
        if coords:
            proj["lat"], proj["lon"] = coords

    if proj.get("lat") and proj.get("lon") and day.get("date"):
        w = _fetch_day_weather(proj["lat"], proj["lon"], day["date"].isoformat())
        if w:
            quart["temp_am"] = w["temp_am"]
            quart["temp_pm"] = w["temp_pm"]
            quart["conditions"] = list(w["conditions"]) if w["conditions"] else []
            return True
        return False
    return False

def _fill_week_weather(proj, jours):
    if not (proj.get("lat") and proj.get("lon")):
        return
    for jour in JOURS:
        day = jours[jour]
        quart = day["quarts"][_day_quart_names(day)[0]]
        if day.get("date") and not quart.get("temp_am"):
            w = _fetch_day_weather(proj["lat"], proj["lon"], day["date"].isoformat())
            if w:
                quart["temp_am"] = w["temp_am"]
                quart["temp_pm"] = w["temp_pm"]
                quart["conditions"] = w["conditions"]

# --------------------------------------------------------------------------
# Logic Métier & Grille
# --------------------------------------------------------------------------
def _roster(quart):
    return ([(n, "P") for n in quart.get("personnel", [])]
            + [(e, "E") for e in quart.get("equipements", [])])

def _to_hours(v):
    """Convertit une valeur de cellule (str/float/None) en float heures (0 si invalide)."""
    try:
        return float(v or 0.0)
    except (TypeError, ValueError):
        return 0.0

def _norm_pair(pair):
    """Normalise une valeur d'heures en couple {'TR','TS'} de float.

    Tolère l'ancien format scalaire ({activité: heures}) — une valeur unique est
    traitée comme du temps régulier — afin qu'un état hérité (ex. session_state
    rémanent après un rechargement de code) ne fasse pas planter la lecture.
    """
    if isinstance(pair, dict):
        return {"TR": _to_hours(pair.get("TR")), "TS": _to_hours(pair.get("TS"))}
    return {"TR": _to_hours(pair), "TS": 0.0}

def _hhmm_to_min(s):
    """'HH:MM' -> minutes depuis minuit ; None si invalide."""
    try:
        h, m = str(s).split(":")
        h, m = int(h), int(m)
        if 0 <= h <= 23 and 0 <= m <= 59:
            return h * 60 + m
    except (ValueError, AttributeError):
        pass
    return None


def _min_to_hhmm(mins):
    """minutes depuis minuit -> 'HH:MM' (borné 00:00..23:59)."""
    mins = max(0, min(int(mins), 23 * 60 + 59))
    return f"{mins // 60:02d}:{mins % 60:02d}"


def _range_hours(debut, fin):
    """Durée en heures décimales entre deux 'HH:MM' ; 0.0 si fin<=début ou invalide."""
    a, b = _hhmm_to_min(debut), _hhmm_to_min(fin)
    if a is None or b is None or b <= a:
        return 0.0
    return (b - a) / 60.0


def _ranges_to_pair(ranges):
    """{'TR': Σ durées TR, 'TS': Σ durées TS} pour une liste de plages."""
    tr = ts = 0.0
    for r in ranges or []:
        d = _range_hours((r or {}).get("debut"), (r or {}).get("fin"))
        if (r or {}).get("type") == "TS":
            ts += d
        else:
            tr += d
    return {"TR": tr, "TS": ts}


def _norm_entry(entry):
    """Normalise une entrée d'heures d'activité -> {'mode','ranges','TR','TS'}.

    Tolère l'ancien format ({'TR','TS'} ou scalaire hérité) -> mode 'direct'.
    En mode 'plage', TR/TS sont dérivés des plages.
    """
    if isinstance(entry, dict):
        ranges = list(entry.get("ranges") or [])
        mode = entry.get("mode") or ("plage" if ranges else "direct")
        if mode == "plage":
            pair = _ranges_to_pair(ranges)
        else:
            pair = {"TR": _to_hours(entry.get("TR")), "TS": _to_hours(entry.get("TS"))}
        return {"mode": mode, "ranges": ranges, "TR": pair["TR"], "TS": pair["TS"]}
    return {"mode": "direct", "ranges": [], "TR": _to_hours(entry), "TS": 0.0}

def _pair_total(pair):
    """Total d'un couple {'TR','TS'} -> float (0 si vide/invalide)."""
    p = _norm_pair(pair)
    return p["TR"] + p["TS"]

def _resource_total(quart, name):
    return float(sum(_pair_total(p) for p in quart["heures"].get(name, {}).values()))

def _quart_activities(quart):
    """Union triée des activités présentes (clés de heures, tous employés)."""
    acts = {a for acts in (quart.get("heures") or {}).values() for a in acts}
    return sorted(acts)

def _quart_total(quart):
    return float(sum(_resource_total(quart, r) for r in quart["heures"]))

def _day_total(day):
    return float(sum(_quart_total(q) for q in day["quarts"].values()))

# --------------------------------------------------------------------------
# Export Excel
# --------------------------------------------------------------------------
def _legacy_day(quart):
    acts = _quart_activities(quart)[:8]
    autres = list(quart["autres"])[:4]
    headers = {f"h{i}": "" for i in range(8)}
    headers.update({f"a{i}": "" for i in range(4)})
    label_to_key = {}
    for i, lbl in enumerate(acts):
        headers[f"h{i}"] = lbl
        label_to_key[lbl] = f"h{i}"
    for i, lbl in enumerate(autres):
        headers[f"a{i}"] = lbl
        label_to_key[lbl] = f"a{i}"

    def build_df(resources, label_col, with_equip=False):
        recs = []
        for name in resources:
            h = quart["heures"].get(name, {})
            rec = {label_col: name}
            for k in HOUR_KEYS:
                rec[k] = None
            for label, key in label_to_key.items():
                if label in h:
                    rec[key] = _pair_total(h[label])
            rec["TR"] = float(sum(_norm_pair(p)["TR"] for p in h.values()))
            rec["TS"] = float(sum(_norm_pair(p)["TS"] for p in h.values()))
            if with_equip:
                rec["Hrs Éq."] = quart["equip_hours"].get(name)
                rec["Code Éq."] = ", ".join(quart["equip_codes"].get(name, []))
            rec["Prime"] = quart["prime"].get(name)
            rec["Commentaire"] = quart["commentaire_ligne"].get(name, "")
            recs.append(rec)
        cols = [label_col] + HOUR_KEYS + ["TR", "TS"]
        if with_equip:
            cols += ["Hrs Éq.", "Code Éq."]
        cols += ["Prime", "Commentaire"]
        return pd.DataFrame(recs, columns=cols)

    return {
        "description": quart.get("description", ""),
        "responsable": quart.get("responsable", ""),
        "temp_am": quart.get("temp_am"), "temp_pm": quart.get("temp_pm"),
        "conditions": quart.get("conditions", []), "headers": headers,
        "pers": build_df(quart.get("personnel", []), "Nom", with_equip=True),
        "equip": build_df(quart.get("equipements", []), "Véhicule"),
    }

def _add_logo(ws, height_px=58):
    if not os.path.exists(LOGO_PATH): return
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    img = XLImage(LOGO_PATH)
    w = round(292 / 280 * height_px)
    marker = AnchorMarker(col=0, colOff=pixels_to_EMU(12), row=0, rowOff=pixels_to_EMU(8))
    img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(pixels_to_EMU(w), pixels_to_EMU(height_px)))
    ws.add_image(img)

def _build_synthese(ws, proj, legacy_jours, exported_by=""):
    ws.title = "Synthèse"
    _add_logo(ws)
    ws.merge_cells("B1:F1")
    t = ws["B1"]; t.value = "RAPPORT JOURNALIER — ONDEL"; t.font = _F_TITLE; t.fill = _FILL_TITLE
    ws.column_dimensions["A"].width = 20
    # Estampille de l'exportateur — DOIT rester la dernière écriture de la fonction (ws.max_row est évalué ici).
    last = ws.max_row + 2
    cell = ws.cell(row=last, column=1, value=f"Exporté par {exported_by or '—'}")
    cell.font = Font(name="Calibri", size=8, italic=True, color="6B7B7E")

def build_workbook():
    wb = Workbook()
    legacy = {(j, q): _legacy_day(st.session_state.jours[j]["quarts"][q])
              for j in JOURS for q in _day_quart_names(st.session_state.jours[j])}
    _build_synthese(wb.active, st.session_state.projet, legacy, current_user()["name"])
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# --------------------------------------------------------------------------
# UI Components & CSS
# --------------------------------------------------------------------------
def get_css():
    return f"""
    <style>
    /* Global Styles */
    .stApp {{ background-color: {ONDEL_LIGHT_BG}; }}
    /* Réduit le padding par défaut (~6rem haut, ~5rem latéral) de l'app */
    .stMainBlockContainer, .block-container {{
        padding-top: 3rem !important;
        padding-left: 1rem !important;
        padding-right: 1rem !important;
    }}

    /* Header — bannière teal construite comme un conteneur Streamlit stylé,
       pour pouvoir y intégrer le bouton Retour (widget) à gauche. */
    .st-key-ondel_header {{
        background: linear-gradient(90deg, {ONDEL_GREEN} 0%, {ONDEL_GREEN_DARK} 100%);
        border-radius: 12px; padding: 0.6rem 1.25rem;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
        position: relative;
    }}
    /* neutralise les marges par défaut des widgets dans la bannière (évite le débordement) */
    .st-key-ondel_header [data-testid="stElementContainer"],
    .st-key-ondel_header [data-testid="stHorizontalBlock"] {{ margin: 0 !important; }}
    /* hauteur de rangée garantie pour contenir le logo, contenu centré, sans débordement */
    .st-key-ondel_header [data-testid="stHorizontalBlock"] {{
        align-items: center !important; flex-wrap: nowrap !important;
        gap: 0.5rem !important; min-height: 34px !important;
    }}
    /* centre verticalement le contenu DANS chaque colonne (bouton / titre / logo) */
    .st-key-ondel_header [data-testid="stColumn"],
    .st-key-ondel_header [data-testid="column"] {{
        justify-content: center !important; display: flex !important;
        flex-direction: column !important; min-width: 0 !important;
        align-self: stretch !important;
    }}
    /* titre : centré horizontalement ET verticalement sur la bannière (position absolue
       pour éviter le bug des colonnes Streamlit qui collapsent en hauteur) */
    .ondel-title {{ color: #fff; font-size: 1.2rem;
        letter-spacing: 1px; font-weight: 700; line-height: 1.2; margin: 0 !important;
        position: absolute; left: 50%; top: 50%;
        transform: translate(-50%, -50%); white-space: nowrap; z-index: 1; }}
    /* logo : dans le flux, à droite du cluster utilisateur (voir coin droit ci-dessous) */
    .logo-wrap {{ display: flex; align-items: center; }}
    .logo-chip {{ background: #fff; border-radius: 7px; padding: 4px 9px;
        display: inline-flex; align-items: center; }}
    .logo-chip img {{ height: 32px; display: block; }}
    /* Bouton Retour fondu dans la bannière */
    .st-key-hdr_retour button {{
        background: rgba(255,255,255,0.15) !important; color: #fff !important;
        border: 1px solid rgba(255,255,255,0.45) !important; padding: 0.25rem 0.7rem !important;
    }}
    .st-key-hdr_retour button:hover {{
        background: rgba(255,255,255,0.28) !important; border-color: #fff !important; color: #fff !important;
    }}
    /* Coin droit : logo + identité connectée + déconnexion, sur UNE ligne,
       alignés à droite (le bloc vertical Streamlit est forcé en ligne). */
    .st-key-ondel_header [data-testid="stColumn"]:last-child [data-testid="stVerticalBlock"] {{
        flex-direction: row !important; flex-wrap: nowrap !important;
        align-items: center !important; justify-content: flex-end !important;
        gap: 0.6rem !important;
    }}
    .st-key-ondel_header [data-testid="stColumn"]:last-child [data-testid="stElementContainer"] {{
        width: auto !important; flex: 0 0 auto !important;
    }}
    /* Identité connectée + déconnexion, alignées à droite sur la ligne du
       titre « Tableau de bord hebdomadaire » (fond clair). */
    .st-key-dash_user [data-testid="stVerticalBlock"] {{
        flex-direction: row !important; flex-wrap: nowrap !important;
        align-items: center !important; justify-content: flex-end !important;
        gap: 0.6rem !important;
    }}
    .st-key-dash_user [data-testid="stElementContainer"] {{
        width: auto !important; flex: 0 0 auto !important;
    }}
    .dash-user-name {{ color: {ONDEL_ACCENT}; font-weight: 600; font-size: 0.9rem;
        white-space: nowrap; }}
    .st-key-hdr_logout button {{
        background: #fff !important; color: {ONDEL_GREEN_DARK} !important;
        border: 1px solid {ONDEL_GREEN} !important;
        padding: 0.12rem 0.75rem !important; font-size: 0.8rem !important;
        min-height: 0 !important; height: auto !important; border-radius: 6px !important;
        line-height: 1.4 !important; white-space: nowrap;
    }}
    .st-key-hdr_logout button:hover {{
        background: {ONDEL_LIGHT_BG} !important; border-color: {ONDEL_GREEN_DARK} !important;
        color: {ONDEL_GREEN_DARK} !important;
    }}

    /* Buttons */
    .stButton > button {{ 
        border-radius: 8px; font-weight: 600; 
        transition: all 0.2s ease;
    }}
    
    /* Day Cards Dashboard */
    div[class*="st-key-go_"] button {{
        min-height: 100px !important; height: 100% !important;
        white-space: normal !important; text-align: left !important;
        justify-content: flex-start !important; align-items: flex-start !important;
        padding: 16px !important; border-radius: 12px !important;
        border: 1px solid #e2e8f0 !important; background: #FFFFFF !important;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05) !important;
        position: relative !important;
    }}
    div[class*="st-key-go_"] button:hover {{ 
        border-color: {ONDEL_GREEN} !important; 
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0,0,0,0.08) !important;
    }}
    
    /* Resource Row */
    .resource-container {{
        background: white; border-radius: 10px; padding: 12px;
        margin-bottom: 10px; border: 1px solid #edf2f7;
    }}
    .resource-header {{
        display: flex; justify-content: space-between; align-items: center;
        margin-bottom: 8px; border-bottom: 1px solid #f7fafc; padding-bottom: 5px;
    }}
    
    /* Weather Widget */
    .weather-box {{
        background: #f1f5f9; border-radius: 8px; padding: 10px; text-align: center;
    }}
    
    /* === Navigation jour (nav_card) ===
       Card blanche avec ombre légère, boutons prev/next stylisés tablette,
       date en deux lignes (jour gras + date complète en secondaire),
       bouton Copier secondaire centré et plus discret. */
    .st-key-nav_card {{
        background: #FFFFFF;
        border: 1px solid #E2E8F0;
        border-radius: 12px;
        padding: 12px 16px 10px;
        margin-bottom: 1rem;
        box-shadow: 0 1px 4px rgba(15, 23, 42, 0.06);
    }}
    /* Supprime le divider Streamlit en dessous — la card fait déjà la séparation */
    .st-key-nav_card + [data-testid="stDivider"] {{ display: none !important; }}

    /* Boutons prev/next : style ghost teal, touch target 48px, icône chevron proéminent */
    .st-key-nav_prev_Lundi button, .st-key-nav_prev_Mardi button,
    .st-key-nav_prev_Mercredi button, .st-key-nav_prev_Jeudi button,
    .st-key-nav_prev_Vendredi button, .st-key-nav_prev_Samedi button,
    .st-key-nav_prev_Dimanche button,
    .st-key-nav_next_Lundi button, .st-key-nav_next_Mardi button,
    .st-key-nav_next_Mercredi button, .st-key-nav_next_Jeudi button,
    .st-key-nav_next_Vendredi button, .st-key-nav_next_Samedi button,
    .st-key-nav_next_Dimanche button {{
        background: transparent !important;
        border: 1.5px solid {ONDEL_GREEN} !important;
        color: {ONDEL_GREEN} !important;
        font-weight: 700 !important;
        font-size: 1rem !important;
        min-height: 48px !important;
        border-radius: 10px !important;
    }}
    .st-key-nav_prev_Lundi button:hover, .st-key-nav_prev_Mardi button:hover,
    .st-key-nav_prev_Mercredi button:hover, .st-key-nav_prev_Jeudi button:hover,
    .st-key-nav_prev_Vendredi button:hover, .st-key-nav_prev_Samedi button:hover,
    .st-key-nav_prev_Dimanche button:hover,
    .st-key-nav_next_Lundi button:hover, .st-key-nav_next_Mardi button:hover,
    .st-key-nav_next_Mercredi button:hover, .st-key-nav_next_Jeudi button:hover,
    .st-key-nav_next_Vendredi button:hover, .st-key-nav_next_Samedi button:hover,
    .st-key-nav_next_Dimanche button:hover {{
        background: {ONDEL_GREEN} !important;
        color: #FFFFFF !important;
    }}
    .st-key-nav_prev_Lundi button:disabled, .st-key-nav_prev_Mardi button:disabled,
    .st-key-nav_prev_Mercredi button:disabled, .st-key-nav_prev_Jeudi button:disabled,
    .st-key-nav_prev_Vendredi button:disabled, .st-key-nav_prev_Samedi button:disabled,
    .st-key-nav_prev_Dimanche button:disabled,
    .st-key-nav_next_Lundi button:disabled, .st-key-nav_next_Mardi button:disabled,
    .st-key-nav_next_Mercredi button:disabled, .st-key-nav_next_Jeudi button:disabled,
    .st-key-nav_next_Vendredi button:disabled, .st-key-nav_next_Samedi button:disabled,
    .st-key-nav_next_Dimanche button:disabled {{
        border-color: #CBD5E1 !important;
        color: #CBD5E1 !important;
    }}

    /* Date au centre : jour en grand + date complète en dessous */
    .nav-date {{
        text-align: center;
        line-height: 1.2;
        padding: 2px 0;
    }}
    .nav-date-day {{
        display: block;
        font-weight: 800;
        font-size: 1.3rem;
        color: {ONDEL_ACCENT};
        letter-spacing: 0.02em;
    }}
    .nav-date-full {{
        display: block;
        font-size: 0.9rem;
        color: #64748B;
        font-weight: 400;
    }}

    /* Bouton Copier : secondaire, petite taille, discret */
    .st-key-copy_Lundi button, .st-key-copy_Mardi button,
    .st-key-copy_Mercredi button, .st-key-copy_Jeudi button,
    .st-key-copy_Vendredi button, .st-key-copy_Samedi button,
    .st-key-copy_Dimanche button {{
        background: transparent !important;
        border: 1px dashed #94A3B8 !important;
        color: #64748B !important;
        font-size: 0.85rem !important;
        font-weight: 500 !important;
        min-height: 36px !important;
        margin-top: 4px !important;
    }}
    .st-key-copy_Lundi button:hover, .st-key-copy_Mardi button:hover,
    .st-key-copy_Mercredi button:hover, .st-key-copy_Jeudi button:hover,
    .st-key-copy_Vendredi button:hover, .st-key-copy_Samedi button:hover,
    .st-key-copy_Dimanche button:hover {{
        background: #F1F5F9 !important;
        border-color: #64748B !important;
        color: {ONDEL_ACCENT} !important;
    }}
    
    /* Total Badge */
    .total-badge {{
        background: {ONDEL_ACCENT}; color: white;
        padding: 4px 12px; border-radius: 20px; font-weight: bold;
    }}

    /* === Carte Météo unique (tablette) ===
       Une seule card avec titre + GPS inline, températures, conditions. */
    .st-key-meteo_card {{
        background: #FFFFFF;
        border: 1px solid #E2E8F0;
        border-radius: 12px;
        padding: 14px 16px;
        margin-bottom: 12px;
        box-shadow: 0 1px 3px rgba(15, 23, 42, 0.04);
    }}

    /* Touch targets tablette : boutons +/- température (compactés à 40px) */
    .st-key-meteo_card [data-testid="stNumberInput"] button {{
        min-width: 40px !important;
        min-height: 40px !important;
        font-size: 1.2rem !important;
    }}
    /* Input température */
    .st-key-meteo_card [data-testid="stNumberInput"] input {{
        min-height: 40px !important;
        font-size: 1rem !important;
    }}
    /* Conditions multiselect : hauteur confort tactile */
    .st-key-meteo_card [data-baseweb="select"] > div:first-child {{
        min-height: 44px !important;
        align-items: center;
        padding: 6px 12px;
    }}
    /* Bouton GPS : compact dans la ligne de titre */
    .st-key-btn_geo_Lundi button,
    .st-key-btn_geo_Mardi button,
    .st-key-btn_geo_Mercredi button,
    .st-key-btn_geo_Jeudi button,
    .st-key-btn_geo_Vendredi button,
    .st-key-btn_geo_Samedi button,
    .st-key-btn_geo_Dimanche button {{
        background: {ONDEL_GREEN} !important;
        color: #FFFFFF !important;
        border: none !important;
        font-weight: 600 !important;
        min-height: 36px !important;
        font-size: 0.875rem !important;
    }}
    .st-key-btn_geo_Lundi button:hover:not(:disabled),
    .st-key-btn_geo_Mardi button:hover:not(:disabled),
    .st-key-btn_geo_Mercredi button:hover:not(:disabled),
    .st-key-btn_geo_Jeudi button:hover:not(:disabled),
    .st-key-btn_geo_Vendredi button:hover:not(:disabled),
    .st-key-btn_geo_Samedi button:hover:not(:disabled),
    .st-key-btn_geo_Dimanche button:hover:not(:disabled) {{
        background: {ONDEL_GREEN_DARK} !important;
    }}
    /* Carte Équipe (au-dessus des activités) — même style de carte */
    .st-key-equipe_box {{
        background: #FFFFFF;
        border: 1px solid #E2E8F0;
        border-radius: 12px;
        padding: 14px 16px;
        margin-bottom: 12px;
        box-shadow: 0 1px 3px rgba(15, 23, 42, 0.04);
    }}
    .st-key-equipe_box > div:first-child p {{
        font-size: 1.05rem !important;
        font-weight: 700 !important;
        color: {ONDEL_ACCENT} !important;
        margin: 0 0 10px 0 !important;
    }}
    /* Touch targets tablette pour les multiselect d'équipe */
    .st-key-equipe_box [data-baseweb="select"] > div:first-child {{
        min-height: 44px !important;
        align-items: center;
        padding: 6px 12px;
    }}
    /* Pills d'équipe (boutons toggle) : grande zone tactile, coins arrondis */
    .st-key-equipe_box [data-testid="stButtonGroup"] button {{
        min-height: 42px !important;
        font-size: 0.95rem !important;
        border-radius: 21px !important;
        margin: 3px 4px 3px 0 !important;
    }}
    /* Pill NON sélectionné : contour discret */
    .st-key-equipe_box button[data-testid="stBaseButton-pills"] {{
        border: 1.5px solid #CBD5E1 !important;
        background: #FFFFFF !important;
        color: {ONDEL_ACCENT} !important;
    }}
    /* Pill SÉLECTIONNÉ : teal plein, texte blanc, contraste fort */
    .st-key-equipe_box button[data-testid="stBaseButton-pillsActive"] {{
        background: {ONDEL_GREEN} !important;
        border: 1.5px solid {ONDEL_GREEN} !important;
        color: #FFFFFF !important;
        font-weight: 600 !important;
    }}
    .st-key-equipe_box button[data-testid="stBaseButton-pillsActive"] p {{
        color: #FFFFFF !important;
    }}
    /* === Titres de sections : même taille et même alignement top ===
       On cible le <p> du titre Activités, et UNIQUEMENT le titre Météo (1ʳᵉ
       colonne de la 1ʳᵉ rangée). L'ancêtre [data-testid="stMarkdown"] est
       décisif : il restreint la règle au vrai titre markdown et EXCLUT les
       labels des number_input (« Température AM/PM »), qui vivent dans
       stWidgetLabel et non dans stMarkdown — sinon « Température AM »
       apparaissait en gras/accent alors que « Température PM » restait normal. */
    .st-key-meteo_card [data-testid="stHorizontalBlock"]:first-of-type [data-testid="stColumn"]:first-child [data-testid="stMarkdown"] p {{
        font-size: 1.05rem !important;
        font-weight: 700 !important;
        color: {ONDEL_ACCENT} !important;
        margin: 0 0 10px 0 !important;
        padding: 0 !important;
        line-height: 1.3 !important;
    }}
    /* Neutralise le padding-top que Streamlit injecte sur la colonne mh1
       (décale le titre météo de ~5px vers le bas vs le titre activités) */
    .st-key-meteo_card [data-testid="stHorizontalBlock"]:first-of-type {{
        margin-top: 0 !important;
        padding-top: 0 !important;
    }}
    .st-key-meteo_card [data-testid="stHorizontalBlock"]:first-of-type [data-testid="stColumn"] {{
        padding-top: 0 !important;
        margin-top: 0 !important;
    }}
    /* Quart de travail — rangée de pastilles tactiles (≥44px), coins arrondis 11px */
    .st-key-quart_box [data-testid="stHorizontalBlock"] {{
        gap: 14px !important;
    }}
    .st-key-quart_box .stButton > button,
    .st-key-quart_box [data-testid="stPopover"] button {{
        min-height: 52px !important;
        border-radius: 11px !important;
        font-weight: 700 !important;
        font-size: 1rem !important;
    }}
    /* Quart actif : bouton sombre plein (accent), texte blanc */
    .st-key-quart_box button[kind="primary"],
    .st-key-quart_box button[data-testid="stBaseButton-primary"] {{
        background: {ONDEL_ACCENT} !important;
        border: 1.5px solid {ONDEL_ACCENT} !important;
        color: #FFFFFF !important;
    }}
    .st-key-quart_box button[kind="primary"]:hover,
    .st-key-quart_box button[data-testid="stBaseButton-primary"]:hover {{
        background: {ONDEL_ACCENT_DK} !important;
        border-color: {ONDEL_ACCENT_DK} !important;
    }}
    /* Quart existant non actif : carte blanche, contour léger, texte sombre */
    .st-key-quart_box button[kind="secondary"],
    .st-key-quart_box button[data-testid="stBaseButton-secondary"] {{
        background: #FFFFFF !important;
        border: 1.5px solid #CBD5E1 !important;
        color: {ONDEL_ACCENT} !important;
    }}
    /* Quart non créé : popover "＋" au style ghost (atténué, pointillé) */
    .st-key-quart_box [data-testid="stPopover"] > div > button {{
        background: #FFFFFF !important;
        border: 1.5px dashed #CBD5E1 !important;
        color: #64748B !important;
        font-weight: 600 !important;
    }}
    /* Bouton "Copier de …" inline dans la rangée de quarts : dashed discret */
    .st-key-quart_box .st-key-copy_Lundi button, .st-key-quart_box .st-key-copy_Mardi button,
    .st-key-quart_box .st-key-copy_Mercredi button, .st-key-quart_box .st-key-copy_Jeudi button,
    .st-key-quart_box .st-key-copy_Vendredi button, .st-key-quart_box .st-key-copy_Samedi button,
    .st-key-quart_box .st-key-copy_Dimanche button {{
        background: #FFFFFF !important;
        border: 1.5px dashed #94A3B8 !important;
        color: #64748B !important;
        font-weight: 600 !important;
    }}

    /* Compteur discret aligné à droite d'un titre de carte ("3 / 7", "2 sélectionné(s)") */
    .count-chip {{ display:block; text-align:right; font-size:0.85rem; font-weight:600; color:#94A3B8; }}
    .count-chip--teal {{ color:{ONDEL_GREEN_DARK}; }}

    /* Conditions météo : pastilles toggle compactes (inactif gris / actif teal) */
    .st-key-meteo_card [data-testid="stButtonGroup"] button {{
        min-height: 34px !important;
        border-radius: 9px !important;
        font-weight: 600 !important;
        font-size: 0.85rem !important;
        padding: 4px 11px !important;
        margin: 2px 4px 2px 0 !important;
    }}
    .st-key-meteo_card button[data-testid="stBaseButton-pills"] {{
        border: 1.5px solid #E2E8F0 !important;
        background: #FFFFFF !important;
        color: #475569 !important;
    }}
    .st-key-meteo_card button[data-testid="stBaseButton-pillsActive"] {{
        background: {ONDEL_GREEN} !important;
        border: 1.5px solid {ONDEL_GREEN} !important;
        color: #FFFFFF !important;
    }}
    .st-key-meteo_card button[data-testid="stBaseButton-pillsActive"] p {{
        color: #FFFFFF !important;
    }}

    /* Équipements : chips sombres en rangée (flex-wrap), clic = retrait (✕) */
    .st-key-equip_chips > [data-testid="stVerticalBlock"] {{
        flex-direction: row !important;
        flex-wrap: wrap !important;
        gap: 8px !important;
        margin-bottom: 12px !important;
    }}
    .st-key-equip_chips [data-testid="stElementContainer"],
    .st-key-equip_chips .stButton {{ width: auto !important; }}
    .st-key-equip_chips .stButton > button {{
        background: {ONDEL_ACCENT} !important;
        color: #FFFFFF !important;
        border: none !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
        padding: 9px 14px !important;
        min-height: auto !important;
    }}
    .st-key-equip_chips .stButton > button:hover {{
        background: {ONDEL_ACCENT_DK} !important;
        color: #FFFFFF !important;
    }}

    /* Saisie des heures — en-tête fixe au-dessus d'un corps défilant.
       L'en-tête (titres de colonnes) reste visible ; les lignes défilent dans
       .st-key-hours_body (hauteur limitée). Approche déterministe (pas de sticky). */
    .st-key-hours_header {{
        background: {ONDEL_LIGHT_BG} !important;
        padding-bottom: 4px !important;
        border-bottom: 1px solid #E2E8F0 !important;
    }}
    .st-key-hours_header [data-testid="stCaptionContainer"] {{ margin-bottom: 0 !important; }}
    .st-key-hours_body {{
        max-height: 60vh !important;
        overflow-y: auto !important;
        scrollbar-gutter: stable !important;   /* garde l'alignement avec l'en-tête */
    }}

    </style>
    """

def _logo_data_uri():
    if not os.path.exists(LOGO_PATH): return ""
    with open(LOGO_PATH, "rb") as f:
        return f"data:image/png;base64,{base64.b64encode(f.read()).decode('ascii')}"

# --------------------------------------------------------------------------
# Views
# --------------------------------------------------------------------------
def view_dashboard():
    t_l, t_r = st.columns([3, 2], vertical_alignment="center")
    t_l.markdown("### 📋 Tableau de bord hebdomadaire")
    with t_r:
        with st.container(key="dash_user"):
            user = current_user()
            st.markdown(
                f'<span class="dash-user-name">👤 {user["name"] or user["email"] or "—"}</span>',
                unsafe_allow_html=True)
            if st.button("Se déconnecter", key="hdr_logout"):
                st.logout()

    with st.container(border=True):
        c1, c2 = st.columns([2, 1])
        proj = st.session_state.projet
        projects = data_source.get_projects()
        
        if not projects:
            c1.error("Impossible de charger les projets.")
            proj["id_project"] = None
        else:
            def _proj_label(no, desc): return f"{no} — {desc}" if desc else no
            labels = [_proj_label(no, desc) for _pid, no, desc in projects]
            by_label = {_proj_label(no, desc): (pid, no) for pid, no, desc in projects}
            current = next((lbl for lbl, (_pid, no) in by_label.items() if no == proj["no"]), None)
            index = labels.index(current) if current in labels else None
            sel = c1.selectbox("Projet", labels, index=index, placeholder="Choisir un projet…")
            if sel: proj["id_project"], proj["no"] = by_label[sel]
            else: proj["id_project"], proj["no"] = None, ""
            
        projet_choisi = bool(proj.get("id_project"))
        
        if projet_choisi and st.session_state.get("staff_prefilled_for") != proj["id_project"]:
            st.session_state["staff_prefilled_for"] = proj["id_project"]

        if "semaine_input" in st.session_state:
            st.session_state["semaine_input"] = _week_start(st.session_state["semaine_input"])
            proj["semaine"] = c2.date_input("Semaine du", key="semaine_input", disabled=not projet_choisi)
        else:
            proj["semaine"] = c2.date_input("Semaine du", value=proj["semaine"], key="semaine_input", disabled=not projet_choisi)
        
        _apply_week_dates(proj["semaine"])

    # Chargement auto du rapport (projet+semaine) depuis Neon, une fois par clé
    if projet_choisi:
        load_report_into_state()

    # Cartes de jours
    today = date.today()
    rules = []
    for jour in JOURS:
        day = st.session_state.jours[jour]
        total = _day_total(day)
        if total > 0:
            rules.append(f".st-key-go_{jour} button {{ background:#f0fdf4 !important; border-color:#bbf7d0 !important; }}")
        if day["date"] == today:
            rules.append(f".st-key-go_{jour} button {{ border:2px solid {ONDEL_GREEN} !important; }}")
            rules.append(f".st-key-go_{jour} button::after {{ content:\"AUJOURD'HUI\"; position:absolute; top:8px; right:12px; font-weight:800; font-size:0.65rem; color:{ONDEL_GREEN}; }}")

    if rules: st.markdown("<style>" + "\n".join(rules) + "</style>", unsafe_allow_html=True)

    cols = st.columns(4)
    for i, jour in enumerate(JOURS):
        day = st.session_state.jours[jour]
        total = _day_total(day)
        date_str = fr_date_short(day["date"]) if day["date"] else ""
        quarts_actifs = [q for q in _day_quart_names(day) if _quart_total(day["quarts"][q]) > 0]
        quarts_str = " · ".join(quarts_actifs)
        with cols[i % 4]:
            label = f"**{jour}**\n\n{date_str}\n\n" + (
                f"✅ {round(total, 1)} h" + (f"\n\n{quarts_str}" if quarts_str else "")
                if total > 0 else "—")
            if st.button(label, key=f"go_{jour}", use_container_width=True, disabled=not projet_choisi):
                st.session_state.active_day = jour
                st.session_state.view = "day_entry"
                _q0 = _day_quart_names(day)[0]
                if not day["quarts"][_q0]["temp_am"] and proj["lat"]:
                    _fill_weather_for_quart(jour, _q0)
                st.rerun()

    st.divider()
    if st.button("📥 EXPORT EXCEL", type="primary", use_container_width=True, disabled=not projet_choisi): 
        st.session_state.view = "export"; st.rerun()

def _add_quart(jour, new_quart, copy_from=None):
    day = st.session_state.jours[jour]
    q = _empty_quart()
    if copy_from and copy_from in day["quarts"]:
        src = day["quarts"][copy_from]
        q["personnel"] = list(src["personnel"])
        q["equipements"] = list(src["equipements"])
        q["responsable"] = src["responsable"]
    day["quarts"][new_quart] = q
    st.session_state[f"_pending_active_quart_{jour}"] = new_quart
    _mark_dirty()


def _render_quart_selector(jour, day, prev_day=None):
    names = _day_quart_names(day)
    pending_key = f"_pending_active_quart_{jour}"
    if pending_key in st.session_state:
        pending = st.session_state.pop(pending_key)
        if pending in names:
            st.session_state[f"active_quart_{jour}"] = pending
    st.session_state.setdefault(f"active_quart_{jour}", names[0])
    current = st.session_state[f"active_quart_{jour}"]

    # Rangée de 3 pastilles fixes (Jour · Soir · Nuit) + bouton « Copier de … » inline.
    # Quart actif = bouton sombre (primary) ; existant = secondaire (clic = bascule) ;
    # non créé = popover "＋" ghost (clic = ajout, copie depuis le quart actif ou vide).
    with st.container(key="quart_box"):
        # Boutons de quart plus larges que le bouton « Copier de … » (poids moindre).
        widths = [1] * len(QUART_NAMES) + ([0.6] if prev_day else [])
        cols = st.columns(widths)
        for col, q in zip(cols, QUART_NAMES):
            if q in names:
                is_active = (q == current)
                if col.button(q, key=f"quart_pick_{jour}_{q}",
                              type="primary" if is_active else "secondary",
                              use_container_width=True) and not is_active:
                    st.session_state[f"active_quart_{jour}"] = q
                    st.rerun()
            else:
                with col.popover(f"＋ {q}", use_container_width=True):
                    st.caption(f"Ajouter le quart **{q}**")
                    if st.button(f"Copier depuis {current}", key=f"copy_quart_{jour}_{q}",
                                 type="primary", use_container_width=True):
                        _add_quart(jour, q, copy_from=current)
                        st.rerun()
                    if st.button("Vide", key=f"empty_quart_{jour}_{q}",
                                 use_container_width=True):
                        _add_quart(jour, q, copy_from=None)
                        st.rerun()
        # Copie « tout le quart » depuis le 1er quart du jour précédent (inline, à droite).
        if prev_day and cols[-1].button(f"📋 Copier de {prev_day}", key=f"copy_{jour}",
                                        use_container_width=True,
                                        help="Copie heures et équipements du jour précédent"):
            prev_day_obj = st.session_state.jours[prev_day]
            src = prev_day_obj["quarts"][_day_quart_names(prev_day_obj)[0]]
            quart = day["quarts"][current]
            quart["heures"] = {r: {a: dict(p) for a, p in acts.items()}
                               for r, acts in src["heures"].items()}
            quart["equip_codes"] = {r: list(c) for r, c in src["equip_codes"].items()}
            quart["equip_hours"] = dict(src["equip_hours"])
            _clear_quart_widget_state(jour, current)
            _mark_dirty()
            st.rerun()


def _current_quart_name(jour):
    day = st.session_state.jours[jour]
    names = _day_quart_names(day)
    key = f"active_quart_{jour}"
    if st.session_state.get(key) not in names:
        st.session_state[key] = names[0]
    return st.session_state[key]


# Poids des colonnes du tableau d'heures par activité (Activité | TR | TS).
_HOURS_COLS = [4, 2, 2]


def _render_ranges_editor(base, initial):
    """Éditeur dynamique de plages (liste de {'debut','fin','type'}).

    Garde une liste à ids stables dans session_state pour que l'ajout/retrait
    ne décale pas les clés de widgets. Renvoie la liste courante des plages.
    """
    lst_key = f"ranges_{base}"
    seq_key = f"rangeseq_{base}"
    if lst_key not in st.session_state:
        seeded = [{"id": i, "debut": (r or {}).get("debut", "08:00"),
                   "fin": (r or {}).get("fin", "08:00"), "type": (r or {}).get("type", "TR")}
                  for i, r in enumerate(initial or [])]
        st.session_state[lst_key] = seeded
        st.session_state[seq_key] = len(seeded)
    rows = st.session_state[lst_key]
    result = []
    for row in rows:
        rid = row["id"]
        dk, fk, kk = f"rg_deb_{base}_{rid}", f"rg_fin_{base}_{rid}", f"rg_knd_{base}_{rid}"
        if dk not in st.session_state:
            _h, _m = (int(x) for x in row["debut"].split(":"))
            st.session_state[dk] = datetime.time(_h, _m)
        if fk not in st.session_state:
            _h, _m = (int(x) for x in row["fin"].split(":"))
            st.session_state[fk] = datetime.time(_h, _m)
        if kk not in st.session_state:
            st.session_state[kk] = row["type"]
        c1, c2, c3, c4, c5 = st.columns([3, 3, 2, 2, 1], vertical_alignment="center")
        deb = c1.time_input("Début", key=dk, step=datetime.timedelta(minutes=15),
                            label_visibility="collapsed", on_change=_mark_dirty)
        fin = c2.time_input("Fin", key=fk, step=datetime.timedelta(minutes=15),
                            label_visibility="collapsed", on_change=_mark_dirty)
        knd = c3.radio("Type", ["TR", "TS"], key=kk, horizontal=True,
                       label_visibility="collapsed", on_change=_mark_dirty)
        deb_s, fin_s = deb.strftime("%H:%M"), fin.strftime("%H:%M")
        dur = _range_hours(deb_s, fin_s)
        c4.markdown(f"**{dur:.2f} h**" if dur > 0 else "⚠️ fin ≤ début")
        if c5.button("✕", key=f"rg_del_{base}_{rid}", help="Retirer la plage"):
            st.session_state[lst_key] = [r for r in rows if r["id"] != rid]
            _mark_dirty()
            st.rerun()
        result.append({"debut": deb_s, "fin": fin_s, "type": knd})
    if st.button("＋ Ajouter une plage", key=f"rg_add_{base}", use_container_width=True):
        nid = st.session_state[seq_key]
        st.session_state[seq_key] = nid + 1
        st.session_state[lst_key] = rows + [{"id": nid, "debut": "08:00", "fin": "08:00", "type": "TR"}]
        _mark_dirty()
        st.rerun()
    return result


def _render_activity_hours(jour, quart_name, name, act, raw):
    """Rend l'éditeur d'heures d'une activité (mode direct ou plage).

    Renvoie l'entrée normalisée {'mode','ranges','TR','TS'}.
    """
    e = _norm_entry(raw)
    base = f"{jour}_{quart_name}_{name}_{act}"
    with st.container(border=True):
        c1, c2 = st.columns([2, 3], vertical_alignment="center")
        c1.markdown(f"**{act}**")
        mode_key = f"mode_{base}"
        if mode_key not in st.session_state:
            st.session_state[mode_key] = "⏱ Plage" if e["mode"] == "plage" else "TR/TS direct"
        mode = c2.radio("Mode de saisie", ["TR/TS direct", "⏱ Plage"], key=mode_key,
                        horizontal=True, label_visibility="collapsed", on_change=_mark_dirty)
        if mode == "⏱ Plage":
            ranges = _render_ranges_editor(base, e["ranges"])
            pair = _ranges_to_pair(ranges)
            st.caption(f"Total : TR {pair['TR']:.2f} h · TS {pair['TS']:.2f} h")
            return {"mode": "plage", "ranges": ranges, "TR": pair["TR"], "TS": pair["TS"]}
        hc1, hc2 = st.columns(2)
        tr_key, ts_key = f"tr_{base}", f"ts_{base}"
        st.session_state.setdefault(tr_key, e["TR"])
        st.session_state.setdefault(ts_key, e["TS"])
        tr = hc1.number_input("TR", key=tr_key, min_value=0.0, step=0.25,
                              format="%.2f", on_change=_mark_dirty)
        ts = hc2.number_input("TS", key=ts_key, min_value=0.0, step=0.25,
                              format="%.2f", on_change=_mark_dirty)
        return {"mode": "direct", "ranges": [], "TR": float(tr), "TS": float(ts)}


def _render_resource_card(jour, quart_name, quart, name, typ, all_activities):
    """Carte de saisie d'une ressource : activités (TR/TS), équipement (employé), prime, commentaire."""
    # --- Activités de la ressource (choisies parmi toutes les activités du projet) ---
    ms_key = f"acts_{jour}_{quart_name}_{name}"
    current_acts = list(quart["heures"].get(name, {}).keys())
    if ms_key not in st.session_state:
        st.session_state[ms_key] = current_acts
    options = sorted(set(all_activities) | set(current_acts))
    sel_acts = st.multiselect("Activités", options, key=ms_key,
                              placeholder="🔍 Activités travaillées…", on_change=_mark_dirty)

    new_heures = {}
    for act in (sel_acts or []):
        entry = _render_activity_hours(jour, quart_name, name, act,
                                       quart["heures"].get(name, {}).get(act, {}))
        if entry["TR"] > 0 or entry["TS"] > 0 or entry["ranges"]:
            new_heures[act] = entry
    if new_heures:
        quart["heures"][name] = new_heures
    elif name in quart["heures"]:
        del quart["heures"][name]

    # --- Équipement rattaché à l'employé (personnel uniquement) ---
    if typ == "P":
        ce1, ce2 = st.columns([3, 1])
        eqc_key = f"eqc_{jour}_{quart_name}_{name}"
        if eqc_key not in st.session_state:
            st.session_state[eqc_key] = list(quart["equip_codes"].get(name, []))
        codes = ce1.pills("Équipement", EQUIP_CODE_VALUES, selection_mode="multi",
                          format_func=_equip_code_label, key=eqc_key, on_change=_mark_dirty)
        if codes:
            quart["equip_codes"][name] = list(codes)
        elif name in quart["equip_codes"]:
            del quart["equip_codes"][name]
        eqh_key = f"eqh_{jour}_{quart_name}_{name}"
        st.session_state.setdefault(eqh_key, _to_hours(quart["equip_hours"].get(name)))
        eqh = ce2.number_input("Hrs Éq.", key=eqh_key, min_value=0.0, step=0.25,
                               format="%.2f", on_change=_mark_dirty)
        if eqh > 0:
            quart["equip_hours"][name] = float(eqh)
        elif name in quart["equip_hours"]:
            del quart["equip_hours"][name]

    # --- Prime + commentaire ---
    cp, cc = st.columns([1, 3])
    p_key = f"p_{jour}_{quart_name}_{name}"
    st.session_state.setdefault(p_key, _to_hours(quart["prime"].get(name)))
    prime = cp.number_input("Prime ($)", key=p_key, min_value=0.0, step=0.5,
                            format="%.2f", on_change=_mark_dirty)
    if prime > 0:
        quart["prime"][name] = float(prime)
    elif name in quart["prime"]:
        del quart["prime"][name]
    c_key = f"c_{jour}_{quart_name}_{name}"
    st.session_state.setdefault(c_key, quart["commentaire_ligne"].get(name, ""))
    com = cc.text_input("Commentaire", key=c_key, on_change=_mark_dirty)
    if com.strip():
        quart["commentaire_ligne"][name] = com.strip()
    elif name in quart["commentaire_ligne"]:
        del quart["commentaire_ligne"][name]


def view_day_entry():
    jour = st.session_state.active_day
    day_idx = JOURS.index(jour)
    day = st.session_state.jours[jour]
    prev_day = JOURS[day_idx - 1] if day_idx > 0 else None
    next_day = JOURS[day_idx + 1] if day_idx < 6 else None

    with st.container(key="nav_card"):
        n1, n2, n3 = st.columns([1, 6, 1], vertical_alignment="center")
        if n1.button(f"◀ {prev_day}" if prev_day else "◀", disabled=not prev_day,
                     use_container_width=True, key=f"nav_prev_{jour}", help="Jour précédent"):
            st.session_state.active_day = prev_day
            st.rerun()
        n2.markdown(
            f'<div class="nav-date">'
            f'<span class="nav-date-day">{jour}</span>'
            f'<span class="nav-date-full">{fr_date_long(day["date"]) if day["date"] else ""}</span>'
            f'</div>', unsafe_allow_html=True)
        if n3.button(f"{next_day} ▶" if next_day else "▶", disabled=not next_day,
                     use_container_width=True, key=f"nav_next_{jour}", help="Jour suivant"):
            st.session_state.active_day = next_day
            st.rerun()

    # Page unique : météo, personnel et saisie des heures sur un seul écran.
    # Le sélecteur de quart est rendu AVANT la résolution du quart courant pour
    # consommer la sélection en attente posée par _add_quart (sinon le quart
    # fraîchement ajouté s'afficherait un cycle en retard).
    _render_quart_selector(jour, day, prev_day)
    quart_name = _current_quart_name(jour)
    quart = day["quarts"][quart_name]
    st.caption(f"👤 Responsable : {current_user()['name'] or '—'}")

    with st.container(border=True, key="meteo_card"):
        header_cols = st.columns([3, 1], vertical_alignment="center")
        header_cols[0].markdown("🌤️ **Météo**")
        proj = st.session_state.projet
        has_date = day.get("date") is not None
        if not has_date:
            st.caption("⚠️ Date non définie")
        _geo_msg = st.session_state.pop(f"geo_msg_{jour}_{quart_name}", None)
        if _geo_msg:
            (st.success if _geo_msg[0] == "success" else st.warning)(_geo_msg[1])
        if header_cols[1].button("📍 GPS", key=f"{jour}_{quart_name}_geo",
                    disabled=not has_date, help="Utiliser ma position GPS actuelle",
                    use_container_width=True):
            st.session_state[f"show_geoloc_{jour}_{quart_name}"] = True
            st.rerun()
        if st.session_state.get(f"show_geoloc_{jour}_{quart_name}"):
            from streamlit_js_eval import get_geolocation
            loc = get_geolocation(component_key=f"geoloc_{jour}_{quart_name}")
            if loc and isinstance(loc, dict) and loc.get("coords"):
                lat = loc["coords"].get("latitude")
                lon = loc["coords"].get("longitude")
                st.session_state.pop(f"show_geoloc_{jour}_{quart_name}", None)
                if lat is not None and lon is not None:
                    proj["lat"] = float(lat); proj["lon"] = float(lon)
                    success = _fill_weather_for_quart(jour, quart_name)
                    st.session_state[f"{jour}_{quart_name}_temp_am"] = quart["temp_am"]
                    st.session_state[f"{jour}_{quart_name}_temp_pm"] = quart["temp_pm"]
                    st.session_state[f"cond_pills_{jour}_{quart_name}"] = [c for c in (quart["conditions"] or []) if c in CONDITIONS]
                    st.session_state[f"geo_msg_{jour}_{quart_name}"] = (
                        ("success", f"Position {proj['lat']:.4f}, {proj['lon']:.4f} — météo récupérée.")
                        if success else
                        ("warning", "Position trouvée mais météo indisponible pour cette date."))
                    st.rerun()
            else:
                st.caption("📡 Autorisez la géolocalisation dans le navigateur…")
        temp_cols = st.columns(2)
        st.session_state.setdefault(f"{jour}_{quart_name}_temp_am", quart["temp_am"])
        st.session_state.setdefault(f"{jour}_{quart_name}_temp_pm", quart["temp_pm"])
        quart["temp_am"] = temp_cols[0].number_input("Température AM (°C)",
                                  key=f"{jour}_{quart_name}_temp_am", step=1.0, format="%.1f")
        quart["temp_pm"] = temp_cols[1].number_input("Température PM (°C)",
                                  key=f"{jour}_{quart_name}_temp_pm", step=1.0, format="%.1f")
        st.caption("Conditions météo")
        cond_key = f"cond_pills_{jour}_{quart_name}"
        if cond_key not in st.session_state:
            st.session_state[cond_key] = [c for c in (quart["conditions"] or []) if c in CONDITIONS]
        _sel_cond = st.pills("Conditions météo", CONDITIONS, selection_mode="multi",
                             key=cond_key, label_visibility="collapsed", on_change=_mark_dirty)
        quart["conditions"] = list(_sel_cond or [])

    with st.container(border=True, key="equipe_box"):
        ph = st.columns([3, 1], vertical_alignment="center")
        ph[0].markdown("👷 **Personnel présent**")
        ph[1].markdown(
            f'<div class="count-chip count-chip--teal">{len(quart.get("personnel", []))} sélectionné(s)</div>',
            unsafe_allow_html=True)
        # Confirmation d'un ajout manuel (posée au run précédent, affichée après le rerun :
        # un st.toast appelé avant st.rerun() serait perdu — même pattern que geo_msg).
        _add_msg_key = f"add_msg_{jour}_{quart_name}"
        if _add_msg_key in st.session_state:
            st.toast(f"✅ « {st.session_state.pop(_add_msg_key)} » ajouté")
        _staff_project = set(data_source.get_project_staff(st.session_state.projet.get("id_project")))
        _staff_all = set(data_source.get_all_staff())
        _staff_current = set(quart.get("personnel", []))
        if _staff_project:
            st.caption("Employés du projet (cliquez pour sélectionner)")
            # Seed unique de l'état (pas de default= avec key= : ça "absorbe"
            # le 1ᵉʳ clic et force un double-clic). Convention du fichier.
            pills_key = f"personnel_pills_{jour}_{quart_name}"
            if pills_key not in st.session_state:
                st.session_state[pills_key] = [e for e in quart.get("personnel", []) if e in _staff_project]
            selected_pills = st.pills(
                "Employés du projet", sorted(_staff_project), selection_mode="multi",
                key=pills_key, label_visibility="collapsed", on_change=_mark_dirty)
            _non_project = [e for e in quart.get("personnel", []) if e not in _staff_project]
            quart["personnel"] = list(selected_pills or []) + _non_project
        _staff_other = sorted(_staff_all - _staff_project - _staff_current)
        if _staff_other:
            st.caption("Ajouter un employé d'un autre projet")
            other_employee = st.selectbox(
                "Autres employés", [""] + _staff_other,
                key=f"other_employee_{jour}_{quart_name}", label_visibility="collapsed",
                placeholder="Rechercher un employé...", index=0)
            if other_employee and other_employee not in quart["personnel"]:
                quart["personnel"].append(other_employee)
                st.rerun()
        st.caption("Ou ajouter manuellement")
        col_input, col_btn = st.columns([4, 1])
        new_emp_key = f"new_employee_{jour}_{quart_name}"
        # Vidage du champ après un ajout : on réinitialise AVANT d'instancier le
        # widget (impossible de modifier sa clé une fois le widget créé dans le run).
        if st.session_state.pop(f"clear_{new_emp_key}", False):
            st.session_state[new_emp_key] = ""
        new_employee = col_input.text_input("Nom", key=new_emp_key,
                                            placeholder="Nom de l'employé...", label_visibility="collapsed")
        if col_btn.button("➕", key=f"add_manual_{jour}_{quart_name}", disabled=not new_employee.strip(),
                            help="Ajouter", use_container_width=True):
            _name = new_employee.strip()
            if _name and _name not in quart["personnel"]:
                quart["personnel"].append(_name)
                st.session_state[_add_msg_key] = _name
                st.session_state[f"clear_{new_emp_key}"] = True
                _mark_dirty()
                st.rerun()

        # --- Saisie des heures (rail + fiche), sur la même carte que le personnel ---
        full_roster = _roster(quart)
        all_activities = data_source.get_activities(st.session_state.projet.get("id_project"))
        st.divider()
        st.markdown("#### 🕐 Saisie des heures")
        if not full_roster:
            st.info("💡 Ajoutez du personnel ci-dessus pour saisir les heures.")
        else:
            # Sélecteur maître-détail : rail de gauche (recherche + liste défilante
            # de boutons) + fiche de saisie à droite. st.button est représentable
            # sous AppTest (contrairement à st.pills), donc testable.
            labels = [n for n, _t in full_roster]
            by_label = {n: (n, t) for n, t in full_roster}
            sel_key = f"resource_sel_{jour}_{quart_name}"
            if st.session_state.get(sel_key) not in labels:
                st.session_state[sel_key] = labels[0]
            col_rail, col_pane = st.columns([1, 2], gap="medium")
            with col_rail:
                q = st.text_input("Rechercher une ressource", key=f"res_search_{jour}_{quart_name}",
                                  placeholder="🔍 Rechercher une ressource…",
                                  label_visibility="collapsed")
                done = sum(1 for n in labels if _resource_total(quart, n) > 0)
                # La recherche filtre l'affichage du rail uniquement ; elle ne touche
                # jamais sel_key (la fiche en cours reste visible même si la ressource
                # sélectionnée est filtrée hors de la liste).
                filt = [n for n in labels if q.casefold() in n.casefold()] if q else labels
                st.caption(f"{len(filt)} résultat(s) · {done} sur {len(labels)} saisies")
                with st.container(height=300):
                    if not filt:
                        st.caption("Aucune ressource ne correspond.")
                    for n in filt:
                        _n, t = by_label[n]
                        tot = _resource_total(quart, n)
                        ic = "👷" if t == "P" else "🚜"
                        status = "🟢" if tot > 0 else "⚪"
                        is_sel = (n == st.session_state[sel_key])
                        if st.button(f"{ic} {n} · {status} {tot:.1f} h",
                                     key=f"pick_{jour}_{quart_name}_{n}",
                                     use_container_width=True,
                                     type="primary" if is_sel else "secondary"):
                            st.session_state[sel_key] = n
                            st.rerun()
            with col_pane:
                name, typ = by_label[st.session_state[sel_key]]
                icon = "👷" if typ == "P" else "🚜"
                st.markdown(f"##### {icon} {name} — {_resource_total(quart, name):.1f} h")
                _render_resource_card(jour, quart_name, quart, name, typ, all_activities)

    quart["description"] = st.text_input("📝 Note du quart", quart["description"],
                                         placeholder="Commentaire sur le quart...",
                                         key=f"note_{jour}_{quart_name}", on_change=_mark_dirty)

    st.divider()
    # Validation douce : on signale ce qui manque sans bloquer l'enregistrement
    # (0/négatif = température remplie ; seul None compte comme « non remplie »).
    missing = []
    if quart["temp_am"] is None and quart["temp_pm"] is None:
        missing.append("une température (AM ou PM)")
    if not quart.get("personnel"):
        missing.append("du personnel")
    sb1, sb2 = st.columns([3, 1], vertical_alignment="center")
    if missing:
        sb1.info("Pour continuer, pensez à ajouter : " + ", ".join(missing) + ".")
    elif st.session_state.get("dirty"):
        sb1.warning("⚠️ Modifications non enregistrées — pensez à enregistrer avant de quitter.")
    else:
        sb1.caption("✓ Toutes les modifications sont enregistrées.")
    if sb2.button("💾 Enregistrer", use_container_width=True, type="primary", key=f"save_{jour}"):
        ok, msg = save_report_from_state()
        (st.success if ok else st.error)(msg)

def view_export():
    st.subheader("📥 Export Excel")
    with st.container(border=True):
        st.write("Le fichier Excel contiendra la synthèse de la semaine ainsi que le détail par jour.")
        if st.button("🚀 Générer le fichier Excel", type="primary", use_container_width=True):
            buf = build_workbook()
            st.download_button("⬇️ Télécharger .xlsx", buf, file_name=f"Rapport_{st.session_state.projet['no']}.xlsx", use_container_width=True)

def main():
    st.set_page_config(page_title="Ondel Rapport journalier", layout="wide", initial_sidebar_state="collapsed")
    if not st.user.is_logged_in:
        st.markdown(get_css(), unsafe_allow_html=True)
        uri = _logo_data_uri()
        with st.container(key="ondel_header"):
            _, c, _ = st.columns([1, 4, 1], vertical_alignment="center")
            c.markdown('<div class="ondel-title">RAPPORTS JOURNALIERS</div>',
                       unsafe_allow_html=True)
        st.markdown(
            f'<div style="text-align:center;margin:2rem 0;">'
            f'<img src="{uri}" style="height:64px;"></div>',
            unsafe_allow_html=True)
        st.info("Accès réservé aux employés ELEM. Connectez-vous avec votre compte Microsoft.")
        _, c, _ = st.columns([1, 2, 1])
        if c.button("🔑 Se connecter avec Microsoft", use_container_width=True,
                    type="primary"):
            st.login()
        st.stop()
    init_state()
    # Crée les tables de rapports au premier rendu (idempotent, une fois par session)
    if not st.session_state.schema_ready:
        try:
            reports.ensure_schema()
            st.session_state.schema_ready = True
        except Exception:
            pass  # BD injoignable : l'app reste utilisable, l'enregistrement signalera l'erreur
    st.markdown(get_css(), unsafe_allow_html=True)
    
    uri = _logo_data_uri()
    with st.container(key="ondel_header"):
        b_l, b_c, b_r = st.columns([1, 4, 1], vertical_alignment="center")
        with b_l:
            if st.session_state.view != "dashboard" and st.button("⬅️ Retour", key="hdr_retour"):
                st.session_state.view = "dashboard"; st.rerun()
        b_c.markdown('<div class="ondel-title">RAPPORTS JOURNALIERS</div>', unsafe_allow_html=True)
        with b_r:
            st.markdown(
                f'<div class="logo-wrap"><span class="logo-chip"><img src="{uri}"></span></div>',
                unsafe_allow_html=True)
    
    if st.session_state.view == "dashboard": view_dashboard()
    elif st.session_state.view == "day_entry": view_day_entry()
    elif st.session_state.view == "export": view_export()

if __name__ == "__main__":
    main()
