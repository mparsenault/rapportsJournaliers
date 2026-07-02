"""Génération Excel des rapports journaliers (présentation).

La mise en forme des DONNÉES reste dans app.py (`_legacy_day`) ; ce module
construit les classeurs openpyxl. Il importe `app` ; app NE DOIT PAS importer
excel_report au niveau module (import paresseux dans les vues) pour éviter un
import circulaire.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import app

# --- Styles Ondel (déplacés depuis app.py) -------------------------------
_TEAL = "0999AA"
_TEAL_DK = "077A88"
_BAND = "EEF8F9"
_GREY = "D9E2E4"
_THIN = Side(style="thin", color=_GREY)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_F_TITLE = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
_F_HEAD = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_F_LABEL = Font(name="Calibri", size=10, bold=True, color=_TEAL_DK)
_F_TOTAL = Font(name="Calibri", size=10, bold=True, color="0E2A2E")
_FILL_TITLE = PatternFill("solid", fgColor=_TEAL)
_FILL_HEAD = PatternFill("solid", fgColor=_TEAL_DK)
_FILL_BAND = PatternFill("solid", fgColor=_BAND)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")
_HOURS_FMT = "0.00"

_BAND_LT = "F5FBFC"          # bande claire secondaire
_SIGN_LINE = "9FB0B2"        # ligne de signature
_F_SUB = Font(name="Calibri", size=9, color="FFFFFF")
_F_SIGN = Font(name="Calibri", size=9, color="5F6E70")
_F_LEGEND = Font(name="Calibri", size=8, color="5F6E70")
_RIGHT = Alignment(horizontal="right", vertical="center")


def _add_logo(ws, height_px=58):
    import os
    if not os.path.exists(app.LOGO_PATH):
        return
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    img = XLImage(app.LOGO_PATH)
    w = round(292 / 280 * height_px)
    marker = AnchorMarker(col=0, colOff=pixels_to_EMU(12), row=0, rowOff=pixels_to_EMU(8))
    img.anchor = OneCellAnchor(_from=marker,
                               ext=XDRPositiveSize2D(pixels_to_EMU(w), pixels_to_EMU(height_px)))
    ws.add_image(img)


def _safe_title(name):
    """Titre de feuille Excel : ≤31 car., sans caractères interdits."""
    bad = '[]:*?/\\'
    return "".join(c for c in str(name) if c not in bad)[:31] or "Feuille"


def _write_row(ws, row, values, *, bold=False, fill=None, fmt=None):
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = _BORDER
        cell.alignment = _LEFT if col == 1 else _CENTER
        if bold:
            cell.font = _F_HEAD if fill is _FILL_HEAD else _F_TOTAL
        if fill is not None:
            cell.fill = fill
        if fmt and isinstance(val, (int, float)):
            cell.number_format = fmt


# Colonnes fixes (regroupement par ressource : activités + plages en sous-lignes).
_PERS_COLS = ["Nom / Activité", "Plage", "TR", "TS", "Hrs Éq.", "Code Éq.", "Prime", "Travaux effectués"]
_EQUIP_COLS = ["Véhicule / Activité", "Plage", "TR", "TS", "Prime", "Commentaire"]
_NCOL = len(_PERS_COLS)  # largeur de la feuille (le tableau personnel est le plus large)


def _col_width(name):
    """Largeur selon le rôle de la colonne."""
    return {
        "Nom / Activité": 40, "Véhicule / Activité": 40, "Plage": 14,
        "TR": 6.5, "TS": 6.5, "Hrs Éq.": 9, "Code Éq.": 13, "Prime": 9,
        "Commentaire": 30, "Travaux effectués": 30,
    }.get(name, 12)


def _apply_widths(ws, cols):
    for i, name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = _col_width(name)


def _banner(ws, row, text):
    """Bande d'info (fusionnée A:dernière colonne), fond teal, texte blanc."""
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=_NCOL)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _F_HEAD
    cell.fill = _FILL_HEAD
    cell.alignment = _LEFT
    ws.row_dimensions[row].height = 20


def _quart_header(ws, row, qname, quart, exported_by=""):
    """Bandeau de quart (teal foncé) : 'Quart X · Resp. Y', puis une sous-ligne
    Température AM/PM + Conditions (bande claire). Renvoie la prochaine ligne."""
    resp = quart.get("responsable") or exported_by
    label = f"Quart {qname}"
    if resp:
        label += f"    ·    Resp. : {resp}"
    _banner(ws, row, label)
    row += 1

    tam, tpm = quart.get("temp_am"), quart.get("temp_pm")
    conds = ", ".join(quart.get("conditions") or [])
    parts = []
    if tam is not None or tpm is not None:
        a = tam if tam is not None else "—"
        p = tpm if tpm is not None else "—"
        parts.append(f"Température : AM {a} / PM {p}")
    if conds:
        parts.append(f"Conditions : {conds}")
    if parts:
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=_NCOL)
        cell = ws.cell(row=row, column=1, value="    ·    ".join(parts))
        cell.font = _F_LABEL
        cell.fill = _FILL_BAND
        cell.alignment = _LEFT
        ws.row_dimensions[row].height = 18
        row += 1
    return row


def _title_band(ws, projet):
    """Bandeau titre teal (lignes 1-2) : logo, 'Rapport journalier' + sous-titre,
    et à droite 'Projet <no>' + pagination. Renvoie la prochaine ligne libre."""
    for r in (1, 2):
        for c in range(1, _NCOL + 1):
            ws.cell(row=r, column=c).fill = _FILL_TITLE
    ws.merge_cells(start_row=1, end_row=1, start_column=2, end_column=_NCOL - 1)
    t = ws.cell(row=1, column=2, value="Rapport journalier")
    t.font = _F_TITLE
    t.alignment = _LEFT
    sub = ws.cell(row=2, column=2, value="Ondel")
    sub.font = _F_SUB
    sub.alignment = _LEFT
    p = ws.cell(row=1, column=_NCOL, value=f"Projet {projet.get('no') or ''}".strip())
    p.font = _F_SUB
    p.alignment = _RIGHT
    pg = ws.cell(row=2, column=_NCOL, value="Page 1 de 1")
    pg.font = _F_SUB
    pg.alignment = _RIGHT
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 16
    _add_logo(ws)
    return 4  # ligne 3 laissée vide, méta démarre en 4


def _meta_block(ws, row, projet, jour_name, day):
    """Bloc méta jour : Date, Adresse. Renvoie la prochaine ligne libre."""
    d = day.get("date")
    date_txt = app.fr_date_long(d) if d else ""
    ws.cell(row=row, column=1, value="Date :").font = _F_LABEL
    ws.cell(row=row, column=2, value=f"{jour_name} {date_txt}".strip())
    row += 1
    ws.cell(row=row, column=1, value="Adresse :").font = _F_LABEL
    ws.cell(row=row, column=2, value=projet.get("adresse") or "")
    return row + 2  # une ligne vide avant le premier quart


def _stamp(ws, exported_by):
    """Estampille discrète en pied de feuille."""
    last = ws.max_row + 2
    cell = ws.cell(row=last, column=1, value=f"Exporté par {exported_by or '—'}")
    cell.font = Font(name="Calibri", size=8, italic=True, color="6B7B7E")


def _quart_hour_totals(quart):
    """Totaux (TR, TS, heures d'équipement) d'un quart, dérivés des heures saisies."""
    tr = ts = 0.0
    for acts in (quart.get("heures") or {}).values():
        for entry in (acts or {}).values():
            norm = app._norm_entry(entry)
            tr += float(norm.get("TR") or 0)
            ts += float(norm.get("TS") or 0)
    eq = sum(float(v) for v in (quart.get("equip_hours") or {}).values() if v is not None)
    return tr, ts, eq


def _day_total_row(ws, row, tr, ts, eq):
    """Ligne 'Total de la journée' (gras, filet teal au-dessus). Prochaine ligne."""
    row += 1  # espace avant le total
    vals = ["Total de la journée", None, tr, ts, (eq or None)] + [None] * (_NCOL - 5)
    _write_row(ws, row, vals, bold=True, fmt=_HOURS_FMT)
    top = Side(style="medium", color=_TEAL)
    for c in range(1, _NCOL + 1):
        ws.cell(row=row, column=c).border = Border(
            left=_THIN, right=_THIN, top=top, bottom=_THIN)
    return row + 1


_EQUIP_LEGEND = ("Codes d'équipement :  BT chariot · C camion · D détecteur · "
                 "É échafaudage · G grue · N nacelle")


def _equip_legend(ws, row):
    """Légende des codes d'équipement (petite police grise). Prochaine ligne."""
    row += 1
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=_NCOL)
    cell = ws.cell(row=row, column=1, value=_EQUIP_LEGEND)
    cell.font = _F_LEGEND
    cell.alignment = _LEFT
    return row + 1


def _comments_block(ws, row):
    """Libellé 'Commentaires / plaintes / suggestions :' + cadre vide (3 lignes)."""
    row += 1
    ws.cell(row=row, column=1,
            value="Commentaires / plaintes / suggestions :").font = _F_LABEL
    row += 1
    ws.merge_cells(start_row=row, end_row=row + 2, start_column=1, end_column=_NCOL)
    dash = Side(style="dashed", color=_GREY)
    box = Border(left=dash, right=dash, top=dash, bottom=dash)
    for r in range(row, row + 3):
        for c in range(1, _NCOL + 1):
            ws.cell(row=r, column=c).border = box
    return row + 3


def _signature_block(ws, row):
    """Deux blocs de signature vides : 'Revu par' et 'Approuvé par'."""
    row += 2  # espace au-dessus des lignes de signature
    half = _NCOL // 2
    line = Border(bottom=Side(style="thin", color=_SIGN_LINE))
    for c in range(1, _NCOL + 1):
        ws.cell(row=row, column=c).border = line
    row += 1
    ws.cell(row=row, column=1, value="Revu par").font = _F_SIGN
    ws.cell(row=row, column=half + 1, value="Approuvé par").font = _F_SIGN
    return row + 1


def _build_day_sheet(ws, projet, jour_name, day, exported_by=""):
    """Écrit le rapport d'une journée dans la feuille `ws`."""
    ws.title = _safe_title(jour_name)
    _apply_widths(ws, _PERS_COLS)

    row = _title_band(ws, projet)
    row = _meta_block(ws, row, projet, jour_name, day)

    day_tr = day_ts = day_eq = 0.0
    for qname in app._day_quart_names(day):
        quart = day["quarts"][qname]
        if app._quart_total(quart) <= 0:
            continue
        row = _quart_header(ws, row, qname, quart, exported_by)
        if quart.get("description"):
            ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=_NCOL)
            ws.cell(row=row, column=1,
                    value=f"Note : {quart['description']}").alignment = _LEFT
            row += 1

        if quart.get("personnel"):
            row = _write_resource_table(ws, row, quart, quart["personnel"],
                                        _PERS_COLS, with_equip=True)
            row += 1
        if quart.get("equipements"):
            row = _write_resource_table(ws, row, quart, quart["equipements"],
                                        _EQUIP_COLS, with_equip=False)
            row += 1
        row += 1

        tr, ts, eq = _quart_hour_totals(quart)
        day_tr += tr
        day_ts += ts
        day_eq += eq

    row = _day_total_row(ws, row, day_tr, day_ts, day_eq)

    row = _equip_legend(ws, row)
    row = _comments_block(ws, row)
    row = _signature_block(ws, row)

    _stamp(ws, exported_by)


def _write_resource_table(ws, row, quart, names, cols, *, with_equip):
    """Tableau groupé par ressource : ligne nom, puis par activité ses plages
    horaires (début–fin, type TR/TS) — ou une ligne unique en mode direct —, et
    une ligne Total (totaux + champs par ressource). Renvoie la prochaine ligne libre.

    Colonnes : [label, Plage, TR, TS, (Hrs Éq., Code Éq.,) Prime, Commentaire]."""
    ncol = len(cols)
    extra = ncol - 4  # colonnes après [label, Plage, TR, TS]
    _write_row(ws, row, cols, bold=True, fill=_FILL_HEAD)
    row += 1

    heures = quart.get("heures") or {}
    prime = quart.get("prime") or {}
    comm = quart.get("commentaire_ligne") or {}
    eqh = quart.get("equip_hours") or {}
    eqc = quart.get("equip_codes") or {}

    for name in names:
        acts = heures.get(name) or {}
        # Ligne du nom (fusionnée sur toute la largeur).
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=ncol)
        nc = ws.cell(row=row, column=1, value=name)
        nc.font = _F_LABEL
        nc.fill = _FILL_BAND
        nc.alignment = _LEFT
        row += 1

        tr_tot = ts_tot = 0.0
        for label in sorted(acts):
            norm = app._norm_entry(acts[label])
            ranges = norm.get("ranges") or []
            if ranges:
                # Activité (sous-ligne) puis une ligne par plage.
                _write_row(ws, row, ["  " + label] + [None] * (ncol - 1))
                row += 1
                for r in ranges:
                    deb = (r or {}).get("debut")
                    fin = (r or {}).get("fin")
                    typ = "TS" if (r or {}).get("type") == "TS" else "TR"
                    dur = app._range_hours(deb, fin)
                    tr = dur if typ == "TR" else 0.0
                    ts = dur if typ == "TS" else 0.0
                    tr_tot += tr
                    ts_tot += ts
                    _write_row(ws, row, [f"      {deb} – {fin}", typ, tr, ts]
                               + [None] * extra, fmt=_HOURS_FMT)
                    row += 1
            else:
                # Mode direct : une seule ligne avec les heures.
                tr = float(norm.get("TR") or 0)
                ts = float(norm.get("TS") or 0)
                tr_tot += tr
                ts_tot += ts
                _write_row(ws, row, ["  " + label, "directe", tr, ts]
                           + [None] * extra, fmt=_HOURS_FMT)
                row += 1

        if with_equip:
            codes = ", ".join(eqc.get(name) or []) or None
            total = ["Total", None, tr_tot, ts_tot, eqh.get(name), codes,
                     prime.get(name), comm.get(name)]
        else:
            total = ["Total", None, tr_tot, ts_tot, prime.get(name), comm.get(name)]
        _write_row(ws, row, total, bold=True, fmt=_HOURS_FMT, fill=_FILL_BAND)
        row += 1
    return row


def build_day_workbook(projet, jour_name, day, exported_by=""):
    wb = Workbook()
    _build_day_sheet(wb.active, projet, jour_name, day, exported_by)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_week_workbook(projet, jours, jours_order, exported_by=""):
    wb = Workbook()
    wb.remove(wb.active)
    for jour_name in jours_order:
        day = jours.get(jour_name) or {}
        if app._day_total(day) <= 0:
            continue
        ws = wb.create_sheet(title=_safe_title(jour_name))
        _build_day_sheet(ws, projet, jour_name, day, exported_by)
    if not wb.sheetnames:                       # aucune journée remplie
        wb.create_sheet(title="Vide")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_day_email(projet, jour_name, day, exported_by=""):
    """Renvoie (subject, html_body, filename, xlsx_bytes) pour une journée."""
    no = str(projet.get("no") or "")
    d = day.get("date")
    date_iso = d.isoformat() if d else "sans-date"
    date_txt = app.fr_date_long(d) if d else ""
    subject = f"Rapport journalier — {no} — {date_txt} ({jour_name})"
    html_body = (
        f"<p>Bonjour,</p>"
        f"<p>Veuillez trouver ci-joint le rapport journalier du projet "
        f"<b>{no}</b> pour le <b>{jour_name} {date_txt}</b>.</p>"
        f"<p style='color:#6B7B7E;font-size:12px'>Exporté par {exported_by or '—'}.</p>"
    )
    filename = f"Rapport_{no}_{date_iso}.xlsx"
    data = build_day_workbook(projet, jour_name, day, exported_by).getvalue()
    return subject, html_body, filename, data
